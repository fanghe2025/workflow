"""
Attachment Processing Utilities

Extracts text content from various attachment types for ML training.
"""

import os
import io
from pathlib import Path
from typing import Dict, Any, Optional, List
import logging

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from pptx import Presentation
except ImportError:
    Presentation = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class AttachmentProcessor:
    """Process various attachment types and extract text content"""

    def __init__(self, max_file_size_mb: int = 10):
        self.max_file_size_mb = max_file_size_mb
        self.max_file_size_bytes = max_file_size_mb * 1024 * 1024

    def process_attachment(
        self,
        file_path: str,
        content: Optional[bytes] = None,
        content_type: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Process an attachment and extract text content.

        Args:
            file_path: Path to the attachment file
            content: Optional file content as bytes
            content_type: Optional MIME type of the file

        Returns:
            Dictionary with extracted text and metadata
        """
        result = {
            "file_path": file_path,
            "content_type": content_type or self._detect_content_type(file_path),
            "text_content": "",
            "processed": False,
            "error": None,
            "file_size": 0,
        }

        try:
            # Read file if content not provided
            if content is None:
                if not os.path.exists(file_path):
                    result["error"] = f"File not found: {file_path}"
                    return result

                file_size = os.path.getsize(file_path)
                if file_size > self.max_file_size_bytes:
                    result["error"] = f"File too large: {file_size / 1024 / 1024:.2f}MB"
                    return result

                with open(file_path, "rb") as f:
                    content = f.read()

            result["file_size"] = len(content)

            # Extract text based on file type
            extension = Path(file_path).suffix.lower()
            result["text_content"] = self._extract_text(
                content, extension, content_type
            )
            result["processed"] = True

        except Exception as e:
            logger.error(f"Error processing attachment {file_path}: {e}")
            result["error"] = str(e)

        return result

    def _detect_content_type(self, file_path: str) -> str:
        """Detect content type from file extension"""
        extension = Path(file_path).suffix.lower()
        content_types = {
            ".pdf": "application/pdf",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".doc": "application/msword",
            ".txt": "text/plain",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xls": "application/vnd.ms-excel",
            ".ppt": "application/vnd.ms-powerpoint",
        }
        return content_types.get(extension, "application/octet-stream")

    def _extract_text(
        self, content: bytes, extension: str, content_type: Optional[str]
    ) -> str:
        """Extract text from file content based on type"""
        text_parts = []

        try:
            if extension == ".pdf" or (content_type and "pdf" in content_type.lower()):
                text_parts.append(self._extract_from_pdf(content))

            elif extension == ".docx" or (
                content_type and "wordprocessingml" in content_type.lower()
            ):
                text_parts.append(self._extract_from_docx(content))

            elif extension in [".txt"] or (
                content_type and "text/plain" in content_type
            ):
                text_parts.append(content.decode("utf-8", errors="ignore"))

            elif extension in [".xlsx", ".xls"] or (
                content_type and "spreadsheet" in content_type.lower()
            ):
                text_parts.append(self._extract_from_excel(content))

            else:
                logger.warning(f"Unsupported file type: {extension} ({content_type})")
                return ""

        except Exception as e:
            logger.error(f"Error extracting text: {e}")
            return ""

        return "\n\n".join(filter(None, text_parts))

    def _extract_from_pdf(self, content: bytes) -> str:
        """Extract text from PDF"""
        if PyPDF2 is None:
            logger.warning("PyPDF2 not installed. Cannot extract PDF text.")
            return ""

        try:
            pdf_file = io.BytesIO(content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text_parts = []

            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
                except Exception as e:
                    logger.warning(
                        f"Error extracting text from PDF page {page_num}: {e}"
                    )

            return "\n\n".join(text_parts)
        except Exception as e:
            logger.error(f"Error reading PDF: {e}")
            return ""

    def _extract_from_docx(self, content: bytes) -> str:
        """Extract text from DOCX"""
        if Document is None:
            logger.warning("python-docx not installed. Cannot extract DOCX text.")
            return ""

        try:
            doc_file = io.BytesIO(content)
            doc = Document(doc_file)
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            return "\n\n".join(paragraphs)
        except Exception as e:
            logger.error(f"Error reading DOCX: {e}")
            return ""

    def _extract_from_excel(self, content: bytes) -> str:
        """Extract text from Excel files"""
        if openpyxl is None:
            logger.warning("openpyxl not installed. Cannot extract Excel text.")
            return ""

        try:
            excel_file = io.BytesIO(content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    if row_text.strip():
                        sheet_text.append(row_text)
                if sheet_text:
                    text_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(sheet_text))

            return "\n\n".join(text_parts)
        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            return ""

    def process_attachments_from_email(
        self, email_data: Dict[str, Any], attachments_dir: str = "data/attachments"
    ) -> List[Dict[str, Any]]:
        """
        Process all attachments from an email.

        Args:
            email_data: Email data dictionary with attachment information
            attachments_dir: Directory to save attachments

        Returns:
            List of processed attachment results
        """
        processed_attachments = []

        if not email_data.get("hasAttachments", False):
            return processed_attachments

        attachments = email_data.get("attachments", [])
        if not attachments:
            return processed_attachments

        os.makedirs(attachments_dir, exist_ok=True)

        for attachment in attachments:
            attachment_id = attachment.get("id", "unknown")
            attachment_name = attachment.get("name", f"attachment_{attachment_id}")

            # Save attachment if content is provided
            attachment_path = os.path.join(attachments_dir, attachment_name)

            # Process attachment
            # Note: In n8n workflow, attachment content should be downloaded first
            result = self.process_attachment(
                file_path=attachment_path, content_type=attachment.get("contentType")
            )

            result["attachment_id"] = attachment_id
            result["attachment_name"] = attachment_name
            processed_attachments.append(result)

        return processed_attachments
